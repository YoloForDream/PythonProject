import json
import pandas as pd
import timedelta
import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.colors import RED, YELLOW, BLUE, BLACK, WHITE
import urllib.request
import time

# import pymongo #if you want to store data in mongodb

root = os.getcwd()


def getHtml(url):
    html = urllib.request.urlopen(url).read()
    return html


def saveHtml(file_name, file_content):
    with open(file_name.replace('/', '_') + ".json", "wb") as f:
        f.write(file_content)


def SaveAsExcelReport(filename, datalist, datetime, checktime):
    if (datalist):
        csv_headers = ['OBJECTID',
                       'Province_State',
                       'Country_Region',
                       'Last_Update',
                       'Lat',
                       'Long_',
                       'Confirmed',
                       'Deaths',
                       'Recovered'
                       ]
        df = pd.DataFrame(datalist)  # use pandas dataframe to got all dicts
        df.index = range(len(df))  # got the index
        filename = datetime + '_' + filename
        df.to_csv(filename, header=csv_headers, index=False, mode='a+')  # save to csv file
        df = pd.read_csv(filename)  # read the CSV file
        cols = ['OBJECTID', 'Province_State', 'Country_Region', 'Lat', 'Long_', 'Confirmed', 'Deaths', 'Recovered',
                'Last_Update']  # setting excel file column
        df = df.loc[:, cols]  # keep the data display in the excel file follow the header format
        df.drop(['Lat', 'Long_'], axis=1, inplace=True)
        df.fillna({'Deaths': 0, 'Recovered': 0}, inplace=True)
        dt = df.fillna(method='bfill', axis=1)
        for i in range(0, dt.shape[0]):
            dt.iloc[i, 6] = checktime
        result_data = dt
        result_data.to_excel(datetime + '_Global_Infected_Num' + '.xlsx', index=False, sheet_name='Data')
        setExcelFormat(datetime + '_Global_Infected_Num' + '.xlsx')
        rows = GotExcelRows(datetime + '_Global_Infected_Num' + '.xlsx')
        DrewChart(datetime + '_Global_Infected_Num' + '.xlsx', rows)
        return True
    else:
        return False


def getdata(filename):
    infectedlocations = []
    with open(filename, 'r') as f:
        data = json.load(f).get('features')
        lenth = len(data)
        for i in range(0, lenth):
            infectedlocations.append(data[i].get('attributes'))
        return infectedlocations


def GotExcelRows(filename):  # got the total coulumns for the current excel file
    path = filename
    df = pd.DataFrame(pd.read_excel(path))
    rows = df.shape[0]
    return rows


def DrewChart(path, chartlocation):  # drew the char  in the report excel file
    location = chartlocation + 4
    location_1 = 'A' + str(location)
    wb = load_workbook(path)
    ws = wb.active
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Infected Confirmed Numbers"
    chart1.y_axis.title = 'Confirmed'
    chart1.x_axis.title = 'Area'
    data = Reference(ws, min_col=4, min_row=1, max_row=chartlocation)  # setting X-RAY
    cats = Reference(ws, min_col=2, min_row=1, max_row=chartlocation)  # setting Y-RAY
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, location_1)
    wb.save(path)


def setExcelFormat(filename):  # setting excel format
    alignment = Alignment(horizontal='center', vertical='center')  # --Setting Excel Format
    thin = Side(border_style="thin", color=BLACK)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_title_font = Font(name='Times New Roman', size=12, bold=True, color=WHITE)  # setting header format
    row_title_fill = PatternFill(fill_type='solid', fgColor=BLUE)
    content_font = Font(name='Times New Roman', size=10, bold=False,
                        color=BLACK)  # setting the cell format withou the header
    content_fill = PatternFill(fill_type='solid', fgColor=YELLOW)  # Setting Excel Format--
    wb = openpyxl.load_workbook(filename)
    ws = wb['Data']
    ws.row_dimensions[1].height = 20
    ws.column_dimensions['A'].width = 16.0
    ws.column_dimensions['B'].width = 16.0
    ws.column_dimensions['C'].width = 18.0
    ws.column_dimensions['D'].width = 16.0
    ws.column_dimensions['E'].width = 16.0
    ws.column_dimensions['F'].width = 16.0
    ws.column_dimensions['G'].width = 18.0
    for row in ws.rows:
        for cell in row:
            cell.alignment = alignment
            if cell.row == 1:
                cell.border = border
                cell.font = row_title_font
                cell.fill = row_title_fill
            else:
                cell.border = border
                cell.font = content_font
                cell.fill = content_fill

    wb.save(filename)


def timeconvert(timeNum):
    timeTemp = float(timeNum / 1000)
    tupTime = time.localtime(timeTemp)
    standTime = time.strftime("%m/%d/%Y %H:%M:%S", tupTime)
    return standTime


def delfile(filename):
    os.remove(filename)


'''
def insert_db(datalist,time):#if you want to store data in mongodb
    client = pymongo.MongoClient(host='localhost', port=27017)
    db = client['2019nCoVDB']
    collection = db['GlobalDailyReport']
    length = len(datalist)
    for i in range(0,length):
         datalist[i]['Last_Update']=time
         collection.insert_one(datalist[i])
    print('Insert To Mongodb Successfully!')
'''
if __name__ == '__main__':
    aurl = "https://services1.arcgis.com/0MSEUqKaxRlEPj5g/arcgis/rest/services/ncov_cases/FeatureServer/1/query?f=json&where=1%3D1&returnGeometry=false&spatialRel=esriSpatialRelIntersects&outFields=*&orderByFields=Confirmed%20desc%2CCountry_Region%20asc%2CProvince_State%20asc&outSR=102100&resultOffset=0&resultRecordCount=250&cacheHint=true"
    html = getHtml(aurl)
    now_date = datetime.datetime.now()
    currentdate = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    saveHtml(currentdate, html)
    newfilename = currentdate + '.json'
    data = getdata(newfilename)
    orignalTime = data[0]['Last_Update']
    checktime = timeconvert(orignalTime)
    if (SaveAsExcelReport('infected', data, currentdate, checktime)):
        print('Create Report Successfully!')
    else:
        print('Fail to Create Report Successfully!')
    os.remove(newfilename)
    os.remove(currentdate + '_infected')
    # insert_db(data,checktime)# if you want to store data in mongodb




